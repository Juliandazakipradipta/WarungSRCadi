from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mysqldb import MySQL
from config import Config
import requests
from datetime import date
import uuid

import os
import urllib.parse


UNIT_DEFAULT_PCS = {
    'pcs': 1,
    'lusin': 12,
    'box': 10,
}

CATEGORIES = [
    'Beras', 'Minyak', 'Gula', 'Tepung', 'Bumbu & Bahan Masak',
    'Mie Instan', 'Makanan Kaleng', 'Susu', 'Air Mineral', 'Minuman',
    'Snack & Biskuit', 'Kebutuhan Rumah Tangga', 'Perawatan Diri',
    'Produk Bayi', 'Rokok', 'Lainnya'
]

UNIT_LIST = [
    'pcs', 'bungkus', 'box', 'botol', 'kaleng', 'karung', 'pack', 'lusin'
]

def konversi_ke_pcs(jumlah, satuan):
    return jumlah * UNIT_DEFAULT_PCS.get(satuan, 1)

def hitung_stok_awal(jumlah, satuan):
    # Hanya konversi untuk satuan yang jelas ke pcs.
    if satuan in UNIT_DEFAULT_PCS:
        return jumlah * UNIT_DEFAULT_PCS[satuan]
    return jumlah

app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(__file__), 'templates'),
            static_folder=os.path.join(os.path.dirname(__file__), 'static'))
app.config.from_object(Config)

mysql = MySQL(app)

# ============================================================
# HELPER: Kirim notifikasi Telegram
# ============================================================
def kirim_telegram(pesan):
    token   = app.config['TELEGRAM_TOKEN']
    chat_id = app.config['TELEGRAM_CHAT_ID']
    if not token or not chat_id:
        return  # Skip kalau belum diisi
    url  = f"https://api.telegram.org/bot{token}/sendMessage"
    data = {'chat_id': chat_id, 'text': pesan, 'parse_mode': 'HTML'}
    try:
        requests.post(url, data=data, timeout=5)
    except Exception:
        pass

# Helper: Cek stok kritis dan kirim notifikasi (Digabung dalam 1 pesan)
def cek_stok_dan_notif(cursor):
    cursor.execute("""
        SELECT nama_produk, stok, stok_minimum, satuan, expired_date,
               CASE
                   WHEN expired_date IS NOT NULL AND expired_date < CURDATE() THEN 'Expired'
                   WHEN expired_date IS NOT NULL AND expired_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY) THEN 'Hampir Kadaluarsa'
               END AS status_expired,
               CASE
                   WHEN stok <= 0 THEN 'Habis'
                   WHEN stok <= stok_minimum THEN 'Kritis'
               END AS status_stok
        FROM produk
        WHERE stok <= stok_minimum
           OR (expired_date IS NOT NULL AND expired_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY))
    """)
    produk_kritis = cursor.fetchall()
    if not produk_kritis:
        return

    expired_items = []
    kritis_items = []

    for p in produk_kritis:
        if p['status_expired']:
            expired_items.append(p)
        elif p['status_stok']:
            kritis_items.append(p)

    pesan_lines = ["⚠️ <b>PERINGATAN PRODUK - Warung SRC Adi</b>\n"]

    if expired_items:
        pesan_lines.append("🗓️ <b>PRODUK KEDALUWARSA / HAMPIR KADALUARSA:</b>")
        for idx, p in enumerate(expired_items, 1):
            exp_str = p['expired_date'].strftime('%d/%m/%Y') if p['expired_date'] else '-'
            pesan_lines.append(f"{idx}. <b>{p['nama_produk']}</b>\n   Stok: {p['stok']} {p['satuan']} | Status: <b>{p['status_expired']}</b> ({exp_str})")
        pesan_lines.append("")

    if kritis_items:
        pesan_lines.append("📉 <b>PRODUK STOK KRITIS:</b>")
        for idx, p in enumerate(kritis_items, 1):
            pesan_lines.append(f"{idx}. <b>{p['nama_produk']}</b>\n   Stok: <b>{p['stok']} {p['satuan']}</b> (Min: {p['stok_minimum']})")
        pesan_lines.append("")

    pesan_lines.append("<i>Segera lakukan restock atau gunakan stok sebelum kadaluarsa!</i>")
    
    pesan = "\n".join(pesan_lines)
    kirim_telegram(pesan)

# Helper: Dapatkan teks struk belanja formatted
def get_struk_text(id_transaksi, cursor):
    cursor.execute("""
        SELECT t.kode_transaksi, t.tanggal_transaksi, p.no_whatsapp,
               t.total_harga, t.metode_pembayaran
        FROM transaksi t
        LEFT JOIN pelanggan p ON t.id_pelanggan = p.id_pelanggan
        WHERE t.id_transaksi = %s
    """, (id_transaksi,))
    transaksi_data = cursor.fetchone()
    if not transaksi_data:
        return None
    
    cursor.execute("""
        SELECT pr.nama_produk, dt.jumlah, dt.harga_satuan, dt.subtotal
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        WHERE dt.id_transaksi = %s
    """, (id_transaksi,))
    detail_items = cursor.fetchall()
    
    pesan_wa = f"""*STRUK PENJUALAN*
🏪 *Warung SRC Adi*

📋 Kode: *{transaksi_data['kode_transaksi']}*
📅 Tanggal: *{transaksi_data['tanggal_transaksi'].strftime('%d/%m/%Y')}*

*─ DETAIL ITEM ─*"""

    for item in detail_items:
        pesan_wa += f"\n• {item['nama_produk']}\n  {item['jumlah']} pcs × Rp {int(item['harga_satuan']):,} = Rp {int(item['subtotal']):,}"
    
    pesan_wa += f"""

━━━━━━━━━━━━━━━━━━━━━━━━
💰 *Total: Rp {int(transaksi_data['total_harga']):,}*
💳 Metode: *{transaksi_data['metode_pembayaran'].title()}*

Terima kasih telah berbelanja! 🙏
━━━━━━━━━━━━━━━━━━━━━━━━"""
    return pesan_wa


# Helper: Kirim pesan WhatsApp menggunakan Fonnte API Gateway
def kirim_whatsapp_fonnte(no_whatsapp, pesan):
    token = app.config['FONNTE_TOKEN']
    if not token or token == 'GantiDenganTokenFonnteAnda':
        return False, "Token Fonnte belum dikonfigurasi di file .env"
    
    # Format nomor WhatsApp agar berstandar internasional (misal: 0812... -> 62812...)
    clean_phone = ''.join(filter(str.isdigit, no_whatsapp.strip()))
    if clean_phone.startswith('0'):
        clean_phone = '62' + clean_phone[1:]
    elif not clean_phone.startswith('62'):
        clean_phone = '62' + clean_phone
        
    url = "https://api.fonnte.com/send"
    headers = {
        "Authorization": token
    }
    data = {
        "target": clean_phone,
        "message": pesan,
        "countryCode": "62"
    }
    try:
        response = requests.post(url, headers=headers, data=data, timeout=10)
        res_json = response.json()
        if res_json.get('status') == True:
            return True, "Pesan berhasil terkirim via Fonnte."
        else:
            return False, f"Fonnte: {res_json.get('reason', 'Gagal mengirim')}"
    except Exception as e:
        return False, f"Koneksi gagal: {str(e)}"

# ============================================================
# ROUTE: Dashboard
# ============================================================
@app.route('/')
def dashboard():
    cur = mysql.connection.cursor()

    # Omzet & transaksi hari ini
    hari_ini = date.today().strftime('%Y-%m-%d')
    cur.execute("""
        SELECT COUNT(*) AS total_trx, COALESCE(SUM(total_harga), 0) AS total_omzet
        FROM transaksi WHERE tanggal_transaksi = %s
    """, (hari_ini,))
    stat_hari = cur.fetchone()

    # Keuntungan bersih hari ini (omzet - modal)
    cur.execute("""
        SELECT COALESCE(SUM(dt.subtotal - (p.harga_beli * dt.jumlah)), 0) AS keuntungan_bersih
        FROM detail_transaksi dt
        JOIN transaksi t ON dt.id_transaksi = t.id_transaksi
        JOIN produk p ON dt.id_produk = p.id_produk
        WHERE t.tanggal_transaksi = %s
    """, (hari_ini,))
    keuntungan_hari = cur.fetchone()['keuntungan_bersih']

    # Total produk
    cur.execute("SELECT COUNT(*) AS total FROM produk")
    total_produk = cur.fetchone()['total']

    # Stok kritis count (hanya yang tidak kadaluarsa/hampir kadaluarsa)
    cur.execute("""
        SELECT COUNT(*) AS kritis FROM produk 
        WHERE stok <= stok_minimum 
          AND (expired_date IS NULL OR expired_date > DATE_ADD(CURDATE(), INTERVAL 7 DAY))
    """)
    stok_kritis_count = cur.fetchone()['kritis']

    # Kadaluarsa count
    cur.execute("""
        SELECT COUNT(*) AS kadaluarsa FROM produk 
        WHERE expired_date IS NOT NULL 
          AND expired_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY)
    """)
    kadaluarsa_count = cur.fetchone()['kadaluarsa']

    # Total pelanggan
    cur.execute("SELECT COUNT(*) AS total FROM pelanggan")
    total_pelanggan = cur.fetchone()['total']

    # Produk stok kritis list (hanya yang tidak kadaluarsa/hampir kadaluarsa)
    cur.execute("""
        SELECT nama_produk, stok, stok_minimum, satuan
        FROM produk
        WHERE stok <= stok_minimum
          AND (expired_date IS NULL OR expired_date > DATE_ADD(CURDATE(), INTERVAL 7 DAY))
        ORDER BY stok ASC
    """)
    produk_kritis = cur.fetchall()

    # Produk kadaluarsa list
    cur.execute("""
        SELECT nama_produk, stok, satuan, expired_date,
               CASE
                   WHEN expired_date < CURDATE() THEN 'Expired'
                   ELSE 'Hampir Kadaluarsa'
               END AS status_expired
        FROM produk
        WHERE expired_date IS NOT NULL
          AND expired_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY)
        ORDER BY expired_date ASC
    """)
    produk_kadaluarsa = cur.fetchall()

    # 10 transaksi terakhir
    cur.execute("""
        SELECT t.kode_transaksi, t.tanggal_transaksi, p.no_whatsapp AS pelanggan_wa,
               t.total_harga, t.metode_pembayaran
        FROM transaksi t
        LEFT JOIN pelanggan p ON t.id_pelanggan = p.id_pelanggan
        ORDER BY t.created_at DESC LIMIT 10
    """)
    transaksi_terakhir = cur.fetchall()

    # Produk terlaris
    cur.execute("""
        SELECT pr.nama_produk, pr.kategori,
               SUM(dt.jumlah) AS total_terjual,
               SUM(dt.subtotal) AS total_pendapatan
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        GROUP BY dt.id_produk
        ORDER BY total_terjual DESC LIMIT 5
    """)
    produk_terlaris = cur.fetchall()

    cur.close()
    return render_template('dashboard.html',
        stat_hari=stat_hari,
        keuntungan_hari=keuntungan_hari,
        total_produk=total_produk,
        stok_kritis_count=stok_kritis_count,
        kadaluarsa_count=kadaluarsa_count,
        total_pelanggan=total_pelanggan,
        produk_kritis=produk_kritis,
        produk_kadaluarsa=produk_kadaluarsa,
        transaksi_terakhir=transaksi_terakhir,
        produk_terlaris=produk_terlaris
    )

@app.route('/test-telegram')
def test_telegram():
    pesan = "🔔 <b>Uji Coba Notifikasi Telegram - Warung SRC Adi</b>\n\nKoneksi berhasil! Bot Anda sudah terhubung dengan aplikasi Warung SRC Adi."
    kirim_telegram(pesan)
    return "Notifikasi uji coba telah dikirim ke Telegram! Silakan periksa chat bot Anda."

@app.route('/kirim-notif-manual')
def kirim_notif_manual():
    cur = mysql.connection.cursor()
    cek_stok_dan_notif(cur)
    cur.close()
    flash('Pemeriksaan selesai. Notifikasi produk kritis/kedaluwarsa telah dikirim ke Telegram (jika ada barang yang memenuhi kriteria).', 'success')
    return redirect(url_for('dashboard'))

# ============================================================
# ROUTE: Produk & Stok
# ============================================================
@app.route('/produk', methods=['GET', 'POST'])
def produk():
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        action = request.form.get('action')

        # Tambah produk baru
        if action == 'tambah':
            nama         = request.form['nama_produk']
            kategori     = request.form['kategori']
            harga_beli   = float(request.form['harga_beli'])
            harga_jual   = float(request.form['harga_jual'])
            stok_input   = int(request.form['stok'])
            stok_min_input = int(request.form['stok_minimum'])
            satuan       = request.form['satuan']
            expired_date = request.form.get('expired_date') or None
            stok         = hitung_stok_awal(stok_input, satuan)
            stok_min     = hitung_stok_awal(stok_min_input, satuan)
            cur.execute("""
                INSERT INTO produk (nama_produk, kategori, harga_beli, harga_jual, stok, stok_minimum, satuan, expired_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (nama, kategori, harga_beli, harga_jual, stok, stok_min, satuan, expired_date))
            mysql.connection.commit()
            flash('Produk berhasil ditambahkan!', 'success')

        # Edit produk yang sudah ada
        elif action == 'edit':
            id_produk    = int(request.form['id_produk'])
            nama         = request.form['nama_produk']
            kategori     = request.form['kategori']
            harga_beli   = float(request.form['harga_beli'])
            harga_jual   = float(request.form['harga_jual'])
            stok_input   = int(request.form['stok'])
            stok_min_input = int(request.form['stok_minimum'])
            satuan       = request.form['satuan']
            expired_date = request.form.get('expired_date') or None
            stok         = hitung_stok_awal(stok_input, satuan)
            stok_min     = hitung_stok_awal(stok_min_input, satuan)
            cur.execute("""
                UPDATE produk
                SET nama_produk = %s, kategori = %s, harga_beli = %s,
                    harga_jual = %s, stok = %s, stok_minimum = %s,
                    satuan = %s, expired_date = %s
                WHERE id_produk = %s
            """, (nama, kategori, harga_beli, harga_jual, stok, stok_min, satuan, expired_date, id_produk))
            mysql.connection.commit()
            flash('Produk berhasil diupdate!', 'success')

        # Restock produk
        elif action == 'restock':
            id_produk = int(request.form['id_produk'])
            jumlah_unit = int(request.form['jumlah_restock'])
            expired_date = request.form.get('expired_date') or None
            cur.execute("SELECT stok, satuan FROM produk WHERE id_produk = %s", (id_produk,))
            produk_data = cur.fetchone()
            jumlah = hitung_stok_awal(jumlah_unit, produk_data['satuan'])
            stok_lama = produk_data['stok']
            if expired_date:
                cur.execute("UPDATE produk SET stok = stok + %s, expired_date = %s WHERE id_produk = %s", (jumlah, expired_date, id_produk))
            else:
                cur.execute("UPDATE produk SET stok = stok + %s WHERE id_produk = %s", (jumlah, id_produk))
            cur.execute("""
                INSERT INTO log_stok (id_produk, jenis, jumlah, stok_sebelum, stok_sesudah, keterangan)
                VALUES (%s, 'masuk', %s, %s, %s, 'Restock manual')
            """, (id_produk, jumlah, stok_lama, stok_lama + jumlah))
            mysql.connection.commit()
            if produk_data['satuan'] in UNIT_DEFAULT_PCS:
                flash(f'Restock berhasil! Stok bertambah {jumlah_unit} {produk_data["satuan"]} ({jumlah} pcs).', 'success')
            else:
                flash(f'Restock berhasil! Stok bertambah {jumlah_unit} {produk_data["satuan"]}.', 'success')

        # Hapus produk
        elif action == 'hapus':
            id_produk = int(request.form['id_produk'])
            cur.execute("SELECT COUNT(*) AS jumlah FROM detail_transaksi WHERE id_produk = %s", (id_produk,))
            terkait = cur.fetchone()['jumlah']
            if terkait > 0:
                flash('Produk tidak bisa dihapus karena sudah dipakai di transaksi. Hapus atau kosongkan riwayat transaksi terlebih dahulu.', 'danger')
            else:
                cur.execute("DELETE FROM log_stok WHERE id_produk = %s", (id_produk,))
                cur.execute("DELETE FROM produk WHERE id_produk = %s", (id_produk,))
                mysql.connection.commit()
                flash('Produk berhasil dihapus.', 'success')

        cur.close()
        return redirect(url_for('produk'))

    kategori_filter = request.args.get('kategori_filter', '')
    where_clause = ''
    params = ()
    if kategori_filter:
        where_clause = 'WHERE p.kategori = %s'
        params = (kategori_filter,)

    # GET: ambil semua produk dengan informasi keuntungan
    cur.execute(f"""
        SELECT p.*,
               (p.harga_jual - p.harga_beli) AS margin_keuntungan,
               COALESCE(SUM(dt.jumlah), 0) AS total_terjual,
               COALESCE(SUM((dt.harga_satuan - p.harga_beli) * dt.jumlah), 0) AS total_keuntungan,
               CASE 
                   WHEN p.stok <= 0 THEN 'Habis'
                   WHEN p.stok <= p.stok_minimum THEN 'Kritis'
                   WHEN p.stok <= p.stok_minimum * 2 THEN 'Rendah'
                   ELSE 'Aman'
               END AS status_stok,
               CASE
                   WHEN p.expired_date IS NOT NULL AND p.expired_date < CURDATE() THEN 'Expired'
                   WHEN p.expired_date IS NOT NULL AND p.expired_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY) THEN 'Hampir Kadaluarsa'
                   ELSE NULL
               END AS status_expired
        FROM produk p
        LEFT JOIN detail_transaksi dt ON p.id_produk = dt.id_produk
        {where_clause}
        GROUP BY p.id_produk
        ORDER BY p.id_produk DESC
    """, params)
    produk_list = cur.fetchall()
    for p in produk_list:
        if p['satuan'] in ['box', 'lusin']:
            factor = UNIT_DEFAULT_PCS.get(p['satuan'], 1)
            unit_count = p['stok'] // factor
            p['display_stok'] = f"{p['stok']} pcs ({unit_count} {p['satuan']})"
        else:
            p['display_stok'] = f"{p['stok']} {p['satuan']}"
    cur.close()
    return render_template('produk.html', produk_list=produk_list,
        kategori_list=CATEGORIES, satuan_list=UNIT_LIST,
        kategori_filter=kategori_filter)

# ============================================================
# ROUTE: Transaksi
# ============================================================
@app.route('/transaksi', methods=['GET', 'POST'])
def transaksi():
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'hapus':
            id_transaksi = int(request.form['id_transaksi'])
            cur.execute("DELETE FROM detail_transaksi WHERE id_transaksi = %s", (id_transaksi,))
            cur.execute("DELETE FROM transaksi WHERE id_transaksi = %s", (id_transaksi,))
            mysql.connection.commit()
            flash('Riwayat transaksi berhasil dihapus.', 'success')
            cur.close()
            return redirect(url_for('transaksi'))

        id_pelanggan     = request.form.get('id_pelanggan') or None
        no_wa_baru      = request.form.get('pelangganSearch', '').strip()
        tanggal          = request.form['tanggal_transaksi']
        metode_bayar     = request.form['metode_pembayaran']
        catatan          = ''

        if not id_pelanggan and no_wa_baru:
            cur.execute("SELECT id_pelanggan FROM pelanggan WHERE no_whatsapp = %s", (no_wa_baru,))
            existing = cur.fetchone()
            if existing:
                id_pelanggan = existing['id_pelanggan']
            else:
                cur.execute("""
                    INSERT INTO pelanggan (nama_pelanggan, no_whatsapp, alamat)
                    VALUES (%s, %s, %s)
                """, ('', no_wa_baru, ''))
                id_pelanggan = cur.lastrowid
                flash(f'Pelanggan baru "{no_wa_baru}" berhasil didaftarkan!', 'success')

        # Validasi: Pastikan pelanggan ada
        if not id_pelanggan:
            flash('Pelanggan harus dipilih atau didaftarkan untuk melakukan transaksi.', 'danger')
            cur.close()
            return redirect(url_for('transaksi'))

        produk_ids       = request.form.getlist('produk_id[]')
        jumlah_arr       = request.form.getlist('jumlah[]')

        kode = 'TRX-' + date.today().strftime('%Y%m%d') + '-' + str(uuid.uuid4())[:4].upper()
        total = 0
        items = []

        for i, pid in enumerate(produk_ids):
            if not pid or not jumlah_arr[i]:
                continue
            pid = int(pid)
            jml = int(jumlah_arr[i])
            cur.execute("SELECT harga_jual, stok FROM produk WHERE id_produk = %s", (pid,))
            pr = cur.fetchone()
            if pr and pr['stok'] >= jml:
                sub = pr['harga_jual'] * jml
                total += sub
                items.append({'id': pid, 'jml': jml, 'harga': pr['harga_jual'], 'sub': sub})

        if items:
            cur.execute("""
                INSERT INTO transaksi (kode_transaksi, id_pelanggan, tanggal_transaksi,
                                       total_harga, status_pembayaran, metode_pembayaran, catatan)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (kode, id_pelanggan, tanggal, total, 'n/a', metode_bayar, catatan))
            id_trx = cur.lastrowid

            for it in items:
                cur.execute("""
                    INSERT INTO detail_transaksi (id_transaksi, id_produk, jumlah, harga_satuan, subtotal)
                    VALUES (%s, %s, %s, %s, %s)
                """, (id_trx, it['id'], it['jml'], it['harga'], it['sub']))

            mysql.connection.commit()
            cek_stok_dan_notif(cur)

            # Ambil detail pelanggan untuk nomor WA
            cur.execute("SELECT no_whatsapp FROM pelanggan WHERE id_pelanggan = %s", (id_pelanggan,))
            cust_data = cur.fetchone()
            
            success_wa = False
            msg_wa = ""
            pesan_wa = ""
            if cust_data and cust_data['no_whatsapp']:
                pesan_wa = get_struk_text(id_trx, cur)
                if pesan_wa:
                    success_wa, msg_wa = kirim_whatsapp_fonnte(cust_data['no_whatsapp'], pesan_wa)
            
            cur.close()
            flash(f'Transaksi {kode} berhasil disimpan!', 'success')
            
            if success_wa:
                flash('Struk belanja telah otomatis dikirim ke WhatsApp pelanggan.', 'success')
                return redirect(url_for('transaksi'))
            elif "Token Fonnte" in msg_wa:
                # Token belum diset, fallback ke redirect manual di frontend
                clean_phone = ''.join(filter(str.isdigit, cust_data['no_whatsapp'].strip()))
                if clean_phone.startswith('0'):
                    clean_phone = '62' + clean_phone[1:]
                elif not clean_phone.startswith('62'):
                    clean_phone = '62' + clean_phone
                pesan_encoded = urllib.parse.quote(pesan_wa)
                whatsapp_url = f"https://api.whatsapp.com/send?phone={clean_phone}&text={pesan_encoded}"
                return redirect(url_for('transaksi', send_wa=whatsapp_url))
            else:
                flash(f'Gagal mengirim WhatsApp otomatis: {msg_wa}', 'warning')
                return redirect(url_for('transaksi'))
        else:
            flash('Tidak ada item valid atau stok tidak mencukupi.', 'danger')
            cur.close()
            return redirect(url_for('transaksi'))

    # GET: ambil data untuk form
    cur.execute("SELECT * FROM pelanggan ORDER BY id_pelanggan DESC")
    pelanggan_list = cur.fetchall()

    cur.execute("SELECT id_produk, nama_produk, harga_jual, stok, satuan, kategori FROM produk WHERE stok > 0 ORDER BY nama_produk")
    produk_list = cur.fetchall()
    for p in produk_list:
        p['unit_factor'] = UNIT_DEFAULT_PCS.get(p['satuan'], 1)

    # Filter tanggal: default hari ini
    tanggal_filter = request.args.get('tanggal', date.today().strftime('%Y-%m-%d'))

    cur.execute("""
        SELECT t.id_transaksi, t.kode_transaksi, t.tanggal_transaksi, p.no_whatsapp AS no_whatsapp,
               t.total_harga, t.metode_pembayaran
        FROM transaksi t
        LEFT JOIN pelanggan p ON t.id_pelanggan = p.id_pelanggan
        WHERE t.tanggal_transaksi = %s
        ORDER BY t.created_at DESC
    """, (tanggal_filter,))
    riwayat = cur.fetchall()
    cur.close()

    return render_template('transaksi.html',
        pelanggan_list=pelanggan_list,
        produk_list=produk_list,
        riwayat=riwayat,
        kategori_list=CATEGORIES,
        today=date.today().strftime('%Y-%m-%d'),
        tanggal_filter=tanggal_filter
    )

# ============================================================
# ROUTE: Pelanggan
# ============================================================
@app.route('/pelanggan', methods=['GET', 'POST'])
def pelanggan():
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'tambah':
            wa   = request.form['no_whatsapp']
            cur.execute("""
                INSERT INTO pelanggan (nama_pelanggan, no_whatsapp, alamat)
                VALUES (%s, %s, %s)
            """, ('', wa, ''))
            mysql.connection.commit()
            flash('Pelanggan berhasil ditambahkan!', 'success')

        elif action == 'hapus':
            id_pel = int(request.form['id_pelanggan'])
            cur.execute("DELETE FROM pelanggan WHERE id_pelanggan = %s", (id_pel,))
            mysql.connection.commit()
            flash('Pelanggan berhasil dihapus.', 'success')

        elif action == 'edit':
            id_pel = int(request.form['id_pelanggan'])
            wa   = request.form['no_whatsapp']
            cur.execute("""
                UPDATE pelanggan SET no_whatsapp = %s
                WHERE id_pelanggan = %s
            """, (wa, id_pel))
            mysql.connection.commit()
            flash('Profil pelanggan berhasil diupdate!', 'success')

        cur.close()
        return redirect(url_for('pelanggan'))

    cur.execute("""
        SELECT p.*, COUNT(t.id_transaksi) AS total_transaksi,
               COALESCE(SUM(t.total_harga), 0) AS total_belanja
        FROM pelanggan p
        LEFT JOIN transaksi t ON p.id_pelanggan = t.id_pelanggan
        GROUP BY p.id_pelanggan
        ORDER BY total_belanja DESC
    """)
    pelanggan_list = cur.fetchall()
    cur.close()
    return render_template('pelanggan.html', pelanggan_list=pelanggan_list)

# ============================================================
# ROUTE: Laporan
# ============================================================
@app.route('/laporan')
def laporan():
    cur = mysql.connection.cursor()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))

    BULAN_ID = {
        'January': 'Januari', 'February': 'Februari', 'March': 'Maret',
        'April': 'April', 'May': 'Mei', 'June': 'Juni',
        'July': 'Juli', 'August': 'Agustus', 'September': 'September',
        'October': 'Oktober', 'November': 'November', 'December': 'Desember'
    }

    cur.execute("""
        SELECT DATE_FORMAT(tanggal_transaksi, '%Y-%m') AS bulan_raw,
               DATE_FORMAT(tanggal_transaksi, '%M %Y') AS bulan,
               COUNT(*) AS total_trx,
               SUM(total_harga) AS total_omzet,
               SUM(dt.subtotal - (p.harga_beli * dt.jumlah)) AS keuntungan_bersih
        FROM transaksi t
        LEFT JOIN detail_transaksi dt ON t.id_transaksi = dt.id_transaksi
        LEFT JOIN produk p ON dt.id_produk = p.id_produk
        GROUP BY bulan_raw ORDER BY bulan_raw DESC LIMIT 12
    """)
    omzet_bulanan = cur.fetchall()
    # Konversi nama bulan ke Bahasa Indonesia
    for row in omzet_bulanan:
        parts = row['bulan'].split(' ', 1)
        row['bulan'] = BULAN_ID.get(parts[0], parts[0]) + ' ' + parts[1]

    cur.execute("""
        SELECT pr.nama_produk, pr.kategori,
               SUM(dt.jumlah) AS total_terjual,
               SUM(dt.subtotal) AS total_pendapatan,
               SUM((dt.harga_satuan - pr.harga_beli) * dt.jumlah) AS total_keuntungan
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        JOIN transaksi t ON dt.id_transaksi = t.id_transaksi
        WHERE DATE_FORMAT(t.tanggal_transaksi, '%%Y-%%m') = %s
        GROUP BY dt.id_produk ORDER BY total_terjual DESC
    """, (bulan,))
    produk_laporan = cur.fetchall()

    # Top 10 Best Sellers
    cur.execute("""
        SELECT pr.nama_produk, pr.kategori,
               SUM(dt.jumlah) AS total_terjual,
               SUM(dt.subtotal) AS total_pendapatan,
               SUM((dt.harga_satuan - pr.harga_beli) * dt.jumlah) AS total_keuntungan
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        JOIN transaksi t ON dt.id_transaksi = t.id_transaksi
        WHERE DATE_FORMAT(t.tanggal_transaksi, '%%Y-%%m') = %s
        GROUP BY dt.id_produk ORDER BY total_terjual DESC LIMIT 10
    """, (bulan,))
    top_sellers = cur.fetchall()

    # Worst Sellers (produk dengan terjual paling sedikit / ada transaksi)
    cur.execute("""
        SELECT pr.nama_produk, pr.kategori,
               SUM(dt.jumlah) AS total_terjual,
               SUM(dt.subtotal) AS total_pendapatan,
               SUM((dt.harga_satuan - pr.harga_beli) * dt.jumlah) AS total_keuntungan
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        JOIN transaksi t ON dt.id_transaksi = t.id_transaksi
        WHERE DATE_FORMAT(t.tanggal_transaksi, '%%Y-%%m') = %s
        GROUP BY dt.id_produk ORDER BY total_terjual ASC LIMIT 10
    """, (bulan,))
    worst_sellers = cur.fetchall()

    cur.close()

    # Konversi bulan_aktif (YYYY-MM) ke label Indonesia, misal "Juni 2026"
    try:
        from datetime import datetime
        dt_bulan = datetime.strptime(bulan, '%Y-%m')
        nama_en = dt_bulan.strftime('%B')  # "June"
        bulan_label = BULAN_ID.get(nama_en, nama_en) + ' ' + dt_bulan.strftime('%Y')  # "Juni 2026"
    except Exception:
        bulan_label = bulan

    return render_template('laporan.html',
        omzet_bulanan=omzet_bulanan,
        produk_laporan=produk_laporan,
        top_sellers=top_sellers,
        worst_sellers=worst_sellers,
        bulan_aktif=bulan,
        bulan_label=bulan_label
    )

@app.route('/laporan/export')
def export_laporan():
    cur = mysql.connection.cursor()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    
    cur.execute("""
        SELECT pr.nama_produk, pr.kategori,
               SUM(dt.jumlah) AS total_terjual,
               SUM(dt.subtotal) AS total_pendapatan,
               SUM((dt.harga_satuan - pr.harga_beli) * dt.jumlah) AS total_keuntungan
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        JOIN transaksi t ON dt.id_transaksi = t.id_transaksi
        WHERE DATE_FORMAT(t.tanggal_transaksi, '%%Y-%%m') = %s
        GROUP BY dt.id_produk ORDER BY total_terjual DESC
    """, (bulan,))
    produk_laporan = cur.fetchall()
    cur.close()

    BULAN_ID = {
        'January': 'Januari', 'February': 'Februari', 'March': 'Maret',
        'April': 'April', 'May': 'Mei', 'June': 'Juni',
        'July': 'Juli', 'August': 'Agustus', 'September': 'September',
        'October': 'Oktober', 'November': 'November', 'December': 'Desember'
    }

    try:
        from datetime import datetime
        dt_bulan = datetime.strptime(bulan, '%Y-%m')
        nama_en = dt_bulan.strftime('%B')
        bulan_label = BULAN_ID.get(nama_en, nama_en) + ' ' + dt_bulan.strftime('%Y')
    except Exception:
        bulan_label = bulan

    import io
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    
    # Tampilkan garis kisi (gridlines)
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Arial", size=15, bold=True, color="1E1B4B")
    subtitle_font = Font(name="Arial", size=11, italic=True, color="6B7280")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    total_border = Border(
        top=Side(style='thin', color='1E1B4B'),
        bottom=Side(style='double', color='1E1B4B')
    )
    
    # Judul Laporan
    ws['A1'] = "LAPORAN PENJUALAN - WARUNG SRC ADI"
    ws['A1'].font = title_font
    ws['A2'] = f"Periode: {bulan_label}"
    ws['A2'].font = subtitle_font
    
    # Headers
    headers = ['No', 'Nama Produk', 'Kategori', 'Total Terjual', 'Total Pendapatan', 'Total Keuntungan']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center" if col_idx == 1 else "left" if col_idx in [2, 3] else "right",
            vertical="center"
        )
        cell.border = thin_border

    rupiah_format = '"Rp"#,##0'
    pcs_format = '#,##0" pcs"'
    
    for row_idx, row_data in enumerate(produk_laporan, 5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=row_data['nama_produk'])
        ws.cell(row=row_idx, column=3, value=row_data['kategori'])
        
        # Konversi ke int/float agar di-serialize sebagai angka murni di Excel (agar rumus SUM berfungsi)
        c4 = ws.cell(row=row_idx, column=4, value=int(row_data['total_terjual'] or 0))
        c4.number_format = pcs_format
        
        c5 = ws.cell(row=row_idx, column=5, value=float(row_data['total_pendapatan'] or 0))
        c5.number_format = rupiah_format
        
        c6 = ws.cell(row=row_idx, column=6, value=float(row_data['total_keuntungan'] or 0))
        c6.number_format = rupiah_format
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=11)
            cell.border = thin_border
            if col_idx >= 4:
                cell.alignment = Alignment(horizontal="right")
                
    # Total row
    tot_row = len(produk_laporan) + 5
    if len(produk_laporan) > 0:
        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=3)
        total_label_cell = ws.cell(row=tot_row, column=1, value="Total Keseluruhan")
        total_label_cell.font = Font(name="Arial", size=11, bold=True)
        total_label_cell.alignment = Alignment(horizontal="right")
        
        for col_idx in range(4, 7):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_letter}5:{col_letter}{tot_row-1})")
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.border = total_border
            if col_idx == 4:
                cell.number_format = pcs_format
            else:
                cell.number_format = rupiah_format
                
        # Terapkan border total untuk kolom 1 s.d 3
        for col_idx in range(1, 4):
            ws.cell(row=tot_row, column=col_idx).border = total_border
            
    # Auto-fit lebar kolom
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Lewati koordinat sel hasil merge untuk menghindari pelebaran kolom No yang berlebihan
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break
            if is_merged:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    response = make_response(out.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=Laporan_Penjualan_{bulan}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

# ROUTE: Struk (Cetak & Kirim WhatsApp)
# ============================================================
@app.route('/struk/<int:id_transaksi>')
def struk(id_transaksi):
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT t.id_transaksi, t.kode_transaksi, t.tanggal_transaksi, p.no_whatsapp,
               t.total_harga, t.metode_pembayaran
        FROM transaksi t
        LEFT JOIN pelanggan p ON t.id_pelanggan = p.id_pelanggan
        WHERE t.id_transaksi = %s
    """, (id_transaksi,))
    transaksi_data = cur.fetchone()
    if not transaksi_data:
        cur.close()
        flash('Transaksi tidak ditemukan.', 'danger')
        return redirect(url_for('transaksi'))
    
    cur.execute("""
        SELECT dt.id_produk, pr.nama_produk, dt.jumlah, dt.harga_satuan, dt.subtotal
        FROM detail_transaksi dt
        JOIN produk pr ON dt.id_produk = pr.id_produk
        WHERE dt.id_transaksi = %s
    """, (id_transaksi,))
    detail_items = cur.fetchall()
    cur.close()
    
    return render_template('struk.html', transaksi=transaksi_data, items=detail_items)

@app.route('/struk/<int:id_transaksi>/kirim-wa', methods=['POST'])
def kirim_struk_wa(id_transaksi):
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.no_whatsapp 
        FROM transaksi t
        LEFT JOIN pelanggan p ON t.id_pelanggan = p.id_pelanggan
        WHERE t.id_transaksi = %s
    """, (id_transaksi,))
    transaksi_data = cur.fetchone()
    
    if not transaksi_data or not transaksi_data['no_whatsapp']:
        cur.close()
        flash('Pelanggan tidak memiliki nomor WhatsApp.', 'danger')
        return redirect(url_for('transaksi'))
    
    pesan_wa = get_struk_text(id_transaksi, cur)
    cur.close()
    
    if not pesan_wa:
        flash('Transaksi tidak ditemukan atau detail kosong.', 'danger')
        return redirect(url_for('transaksi'))
        
    # Selalu redirect manual ke WhatsApp Web/App
    clean_phone = ''.join(filter(str.isdigit, transaksi_data['no_whatsapp'].strip()))
    if clean_phone.startswith('0'):
        clean_phone = '62' + clean_phone[1:]
    elif not clean_phone.startswith('62'):
        clean_phone = '62' + clean_phone
        
    pesan_encoded = urllib.parse.quote(pesan_wa)
    whatsapp_url = f"https://api.whatsapp.com/send?phone={clean_phone}&text={pesan_encoded}"
    return redirect(whatsapp_url)

# ============================================================
if __name__ == '__main__':
    app.run(debug=True)